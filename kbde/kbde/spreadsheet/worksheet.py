from openpyxl import Workbook as BaseWorkbook,\
                     load_workbook
from openpyxl.writer.excel import save_virtual_workbook

import os


class Worksheet:
    NAME = None
    FIELD_LIST = None

    def __init__(self,filename=None,workbook=None):
        if self.NAME is None:
            raise Exception("must define self.NAME")

        if workbook is None and filename is None:
            raise Exception("must provide a workbook or a filename")

        self.worksheet_created = False

        #Load the workbook or get it from file
        if workbook is not None:
            self.workbook = workbook
            self.filename = None
        else:
            self.filename = filename
            self.workbook = self.openWorkbook(self.filename)

        #Get or create this worksheet
        try:
            self.worksheet = self.workbook.get_sheet_by_name(self.NAME)
        except:
            self.worksheet = self.workbook.create_sheet()
            self.worksheet.title = self.NAME
            self.seedWorksheet()
            self.worksheet_created = True

        self.updateColumnCount()
        self.updateRowCount()

        self.field_map = self.getFieldMap()


    def openWorkbook(self,filename):
        """
        Opens or creates a workbook, based on filename
        """
        try:
            workbook = load_workbook(filename)
        except IOError:
            workbook = BaseWorkbook()
        return workbook


    def seedWorksheet(self):
        index = 0
        if self.FIELD_LIST is not None:
            for field in self.FIELD_LIST:
                self.setCell(-1,index,field)
                index += 1

    def updateColumnCount(self):
        row_num = -1
        column_num = 0
        cell = self.getCell(row_num,column_num)
        while cell.value:
            column_num += 1
            cell = self.getCell(row_num,column_num)
        self.column_count = column_num
        return column_num

    def updateRowCount(self):
        column_num = 0
        row_num = 0
        cell = self.getCell(row_num,column_num)
        while cell.value:
            row_num += 1
            cell = self.getCell(row_num,column_num)
        self.row_count = row_num
        return row_num

    def getFieldMap(self):
        """
        Creates a mapping between header names and column numbers
        """
        field_map = {}

        #Get the first row of the worksheet
        header_list = self.getHeaderList()

        if not self.FIELD_LIST:
            i = 0
            for field in header_list:
                field_map[field] = i
                i += 1
            return field_map

        for field in self.FIELD_LIST:
            if field not in header_list:
                raise HeaderException("field {0} not in header".format(field))
            field_map[field] = header_list.index(field)         
        
        return field_map

    def getHeaderList(self):
        """
        Returns a list of the items in the header row
        """
        header_list = []
        row_num = -1
        column_num = 0
        cell = self.getCell(row_num,column_num)
        while cell.value:
            header_list.append(cell.value)
            column_num += 1
            cell = self.getCell(row_num,column_num)
        return header_list

    def set(self,obj):
        """
        Takes an object
        Creates a new row with the columns mapped
        """
        next_row = self.row_count
        for key in obj:
            if key not in self.field_map:
                continue
            column_num = self.field_map[key]
            cell = self.getCell(next_row,column_num)
            cell.value = obj[key]
        self.row_count += 1

    def get(self,row_number):
        row = tuple(self.worksheet.rows)[row_number+1]
        obj = {}
        for field in self.field_map:
            value = row[self.field_map[field]].value
            obj[field] = value
        return obj

    def getAll(self,function_pointer):
        wb = load_workbook(self.filename,read_only=True)
        ws = wb[self.NAME]
        field_indexes = {self.field_map[key]: key for key in self.field_map}
        r = 0
        for row in ws.rows:
            if r == 0:
                r += 1
                continue
            i = -1
            obj = {}
            for cell in row:
                i += 1
                if i not in field_indexes:
                    continue
                key = field_indexes[i]
                obj[key] = cell.value
            function_pointer(obj)


    def getCell(self,row,column):
        """
        Adapts the indexes for cells and columns so that they start indexig at 0, and ignore the header row
        """
        row += 2
        column += 1
        cell = self.worksheet.cell(row=row,column=column)
        return cell

    def setCell(self,row,column,value):
        cell = self.getCell(row,column)
        cell.value = value
        return cell
        
    def save(self):
        self.workbook.save(self.filename)

    class HeaderException(Exception):
        pass
